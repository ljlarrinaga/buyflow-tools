#!/usr/bin/env python3
"""
BuyFlow Report Generator + Validator v3.0
=========================================
Genera reportes .xlsx de comparación de proveedores y los valida automáticamente.

USO COMO GENERADOR (desde Claude o Claude Code):
    from buyflow import generate_report
    generate_report(config, output_path='reporte.xlsx')

USO COMO VALIDADOR:
    python3 buyflow.py validate reporte.xlsx

USO END-TO-END:
    python3 buyflow.py generate config.json [output.xlsx]

ARQUITECTURA DE FILAS:
  El layout se construye dinámicamente según dos flags booleanos:
    - tiene_descripcion (True/False)
    - tiene_link (True/False)
  Más custom_rows opcionales que se insertan después de Provincia.
  Esto cubre todos los casos: normal, FCD, normal-con-link, etc.

REGLAS DE NEGOCIO (embebidas):
  - Proveedores analizados: siempre 8 salvo que se indique otro número.
  - Si no hay referencia del cliente, comparar contra precio promedio de mercado.
  - Siempre mostrar ahorro % contra algo.
  - Seleccionar los proveedores más significativos para mostrar (no incluir
    opciones carísimas sin justificación).
  - Logo BuyFlow siempre incluido.

CONFIG DICT:
{
    "sheet_name": "Requerimiento",
    "requerimiento": "Se solicita...",
    "recomendacion": "Se recomienda...",
    "proveedores_analizados": 8,       # default 8
    "cantidad": 1,
    "precio_label": "Precio c/iva",
    "costo_total_label": "Costo Total",
    "costo_formula": "=+{col}{price_row}*{qty}",
    "tiene_descripcion": True,         # default False
    "tiene_link": True,                # default True (si algún prov tiene link)
    "custom_rows": [                   # optional
        {"label": "Diámetro", "bold": False},
        {"label": "Potencia", "bold": False},
    ],
    "proveedores": [
        {
            "header": "Referencia",     # o "Proveedor A", etc.
            "modelo": "...",
            "descripcion": "...",        # si tiene_descripcion
            "link": "https://...",       # si tiene_link
            "precio": 50000,             # numeric o "A definir"
            "nombre_proveedor": "...",
            "envio": 0,
            "provincia": "Buenos Aires",
            "plazo_entrega": "Inmediata",
            "financiacion": "Transferencia",
            "garantia": "-",
            "recomendado": "X" | "Alternativa" | None,
            "notas": "..." | None,
            "custom": ["20 pulgadas", "130W"],  # si hay custom_rows
        },
    ]
}
"""

import os, sys, subprocess, json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

PURPLE = PatternFill('solid', fgColor='B24EFF')
NO_FILL = PatternFill(fill_type=None)
FONT10 = Font(name='Arial', size=10)
FONT10B = Font(name='Arial', size=10, bold=True)
FONT14 = Font(name='Arial', size=14)
FONT_LINK = Font(name='Arial', size=10, color='0563C1', underline='single')
ACC = Alignment(horizontal='center', vertical='center')
ACCW = Alignment(horizontal='center', vertical='center', wrap_text=True)
MONEY = '_("$ "* #,##0_);_("$ "* \\(#,##0\\);_("$ "* "-"??_);_(@_)'
PCT = '0%'
MED = Side(style='medium')
NONE_SIDE = Side()

LOGO_SRC = '/mnt/project/Copia_de_LogoBF_fondo_blanco.png'
RECALC = '/mnt/skills/public/xlsx/scripts/recalc.py'


def _brd(t=False, b=False):
    return Border(top=MED if t else NONE_SIDE, bottom=MED if b else NONE_SIDE,
                  left=NONE_SIDE, right=NONE_SIDE)

BTB = _brd(True, True)
BT = _brd(True, False)
BB = _brd(False, True)
BN = _brd(False, False)


# ══════════════════════════════════════════════════════════════════════════════
# ROW MAP — Single source of truth
# ══════════════════════════════════════════════════════════════════════════════

def row_map(tiene_descripcion=False, tiene_link=False, custom_rows=None):
    """Build row map dynamically based on which optional rows are present."""
    r = 10
    rm = dict(req=3, rec=5, prov_anal=7, headers=9)
    rm['modelo'] = r; r += 1
    if tiene_descripcion:
        rm['descripcion'] = r; r += 1
    if tiene_link:
        rm['link'] = r; r += 1
    rm['precio'] = r; r += 1
    rm['proveedor'] = r; r += 1
    rm['envio'] = r; r += 1
    rm['costo_total'] = r; r += 1
    rm['ahorro_d'] = r; r += 1
    rm['ahorro_p'] = r; r += 1
    rm['cap_log'] = r; r += 1
    rm['provincia'] = r; r += 1
    if custom_rows:
        for i in range(len(custom_rows)):
            rm[f'custom_{i}'] = r; r += 1
    rm['plazo'] = r; r += 1
    rm['cond_com'] = r; r += 1
    rm['financiacion'] = r; r += 1
    rm['serv_cli'] = r; r += 1
    rm['garantia'] = r; r += 1
    rm['recomendado'] = r; r += 1
    rm['notas'] = r
    return rm


def border_rules(rm):
    return {
        rm['req']: ('top', 'bottom'), rm['rec']: ('top', 'bottom'),
        rm['prov_anal']: ('top', 'bottom'), rm['headers']: ('top', 'bottom'),
        rm['modelo']: ('top',), rm['costo_total']: ('bottom',),
        rm['ahorro_d']: ('top',), rm['ahorro_p']: ('bottom',),
        rm['cap_log']: ('top',), rm['plazo']: ('bottom',),
        rm['cond_com']: ('top',), rm['financiacion']: ('bottom',),
        rm['serv_cli']: ('top',), rm['garantia']: ('bottom',),
        rm['recomendado']: ('top', 'bottom'), rm['notas']: ('top',),
    }


def bold_rows(rm):
    return {rm['costo_total'], rm['ahorro_d'], rm['ahorro_p'],
            rm['cap_log'], rm['provincia'], rm['cond_com'], rm['serv_cli'],
            rm['recomendado']}


def not_bold_rows(rm):
    keys = ['modelo', 'precio', 'proveedor', 'envio', 'plazo', 'financiacion',
            'garantia', 'notas']
    for k in ('descripcion', 'link'):
        if k in rm: keys.append(k)
    return {rm[k] for k in keys if k in rm}


# ══════════════════════════════════════════════════════════════════════════════
# GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def generate_report(config, output_path='/home/claude/report.xlsx'):
    tiene_desc = config.get('tiene_descripcion', False)
    tiene_link = config.get('tiene_link', True)
    custom = config.get('custom_rows', None)
    rm = row_map(tiene_desc, tiene_link, custom)
    br = border_rules(rm)
    provs = config['proveedores']
    n = len(provs)
    last_ci = 2 + n
    last_cl = get_column_letter(last_ci)
    qty = config.get('cantidad', 1)

    wb = Workbook()
    ws = wb.active
    ws.title = config.get('sheet_name', 'Requerimiento')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 5.3
    ws.column_dimensions['B'].width = 35
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 28

    def sc(cell, val=None, font=FONT10, fill=NO_FILL, align=ACC, brd=BN, nf='General'):
        if val is not None: cell.value = val
        cell.font = font; cell.fill = fill; cell.alignment = align
        cell.border = brd; cell.number_format = nf

    def gbrd(row):
        sides = br.get(row, ())
        return _brd('top' in sides, 'bottom' in sides)

    for r in [2, 4, 6, 8]:
        ws.row_dimensions[r].height = 15.75
    ws.row_dimensions[1].height = 40

    # R3: Requerimiento
    ws.row_dimensions[rm['req']].height = config.get('req_height', 120)
    sc(ws[f'B{rm["req"]}'], 'Requerimiento:', FONT14, PURPLE, ACCW, BTB)
    ws.merge_cells(f'C{rm["req"]}:{last_cl}{rm["req"]}')
    sc(ws[f'C{rm["req"]}'], config['requerimiento'], FONT14, NO_FILL, ACCW, BTB)
    for i in range(1, n):
        sc(ws[f'{get_column_letter(3+i)}{rm["req"]}'], font=FONT14, align=ACCW, brd=BTB)

    # R5: Recomendación
    ws.row_dimensions[rm['rec']].height = config.get('rec_height', 140)
    sc(ws[f'B{rm["rec"]}'], 'Recomendación BuyFlow', FONT14, PURPLE, ACCW, BTB)
    ws.merge_cells(f'C{rm["rec"]}:{last_cl}{rm["rec"]}')
    sc(ws[f'C{rm["rec"]}'], config['recomendacion'], FONT14, NO_FILL, ACCW, BTB)
    for i in range(1, n):
        sc(ws[f'{get_column_letter(3+i)}{rm["rec"]}'], font=FONT14, align=ACCW, brd=BTB)

    # R7: Proveedores analizados
    ws.row_dimensions[rm['prov_anal']].height = 15.75
    sc(ws[f'B{rm["prov_anal"]}'], 'Proveedores analizados', FONT10, PURPLE, ACC, BTB)
    sc(ws[f'C{rm["prov_anal"]}'], config.get('proveedores_analizados', 8), FONT10, NO_FILL, ACC, BTB)

    # Headers
    ws.row_dimensions[rm['headers']].height = 20
    sc(ws[f'B{rm["headers"]}'], None, FONT10, PURPLE, ACC, BTB)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["headers"]}'], p['header'], FONT10, PURPLE, ACC, BTB)

    # Modelo
    ws.row_dimensions[rm['modelo']].height = config.get('modelo_height', 42)
    sc(ws[f'B{rm["modelo"]}'], 'Modelo', FONT10, NO_FILL, ACCW, gbrd(rm['modelo']))
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["modelo"]}'], p.get('modelo', '-'), FONT10, NO_FILL, ACCW, gbrd(rm['modelo']))

    # Descripcion
    if tiene_desc:
        ws.row_dimensions[rm['descripcion']].height = config.get('desc_height', 60)
        sc(ws[f'B{rm["descripcion"]}'], 'Descripcion', FONT10, NO_FILL, ACCW, BN)
        for i, p in enumerate(provs):
            sc(ws[f'{get_column_letter(3+i)}{rm["descripcion"]}'], p.get('descripcion', '-'), FONT10, NO_FILL, ACCW, BN)

    # Link
    if tiene_link:
        ws.row_dimensions[rm['link']].height = 14
        sc(ws[f'B{rm["link"]}'], 'Link', FONT10, NO_FILL, ACC, BN)
        for i, p in enumerate(provs):
            col = get_column_letter(3 + i)
            cell = ws[f'{col}{rm["link"]}']
            url = p.get('link')
            if url:
                cell.value = 'Link'; cell.hyperlink = url; cell.font = FONT_LINK
            else:
                sc(cell, '-', FONT10, NO_FILL, ACC, BN)
            cell.alignment = ACC

    # Precio
    ws.row_dimensions[rm['precio']].height = 20
    sc(ws[f'B{rm["precio"]}'], config.get('precio_label', 'Precio c/iva'), FONT10, NO_FILL, ACC, BN)
    for i, p in enumerate(provs):
        col = get_column_letter(3 + i)
        precio = p.get('precio')
        if isinstance(precio, (int, float)):
            sc(ws[f'{col}{rm["precio"]}'], precio, FONT10, NO_FILL, ACC, BN, MONEY)
        elif isinstance(precio, str) and precio.startswith('='):
            sc(ws[f'{col}{rm["precio"]}'], precio, FONT10, NO_FILL, ACC, BN, MONEY)
        else:
            sc(ws[f'{col}{rm["precio"]}'], precio or '-', FONT10, NO_FILL, ACC, BN)

    # Proveedor nombre
    ws.row_dimensions[rm['proveedor']].height = 20
    sc(ws[f'B{rm["proveedor"]}'], 'Nombre del Proveedor', FONT10, NO_FILL, ACC, BN)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["proveedor"]}'], p.get('nombre_proveedor', '-'), FONT10B, NO_FILL, ACC, BN, MONEY)

    # Envío
    ws.row_dimensions[rm['envio']].height = config.get('envio_height', 20)
    sc(ws[f'B{rm["envio"]}'], 'Envio', FONT10, NO_FILL, ACC, BN)
    for i, p in enumerate(provs):
        col = get_column_letter(3 + i)
        envio = p.get('envio', 0)
        if isinstance(envio, (int, float)):
            sc(ws[f'{col}{rm["envio"]}'], envio, FONT10, NO_FILL, ACC, BN, MONEY)
        else:
            sc(ws[f'{col}{rm["envio"]}'], envio, FONT10, NO_FILL, ACCW, BN)

    # Costo Total
    ws.row_dimensions[rm['costo_total']].height = 20
    brd_ct = gbrd(rm['costo_total'])
    sc(ws[f'B{rm["costo_total"]}'], config.get('costo_total_label', 'Costo Total'), FONT10B, NO_FILL, ACC, brd_ct)
    ct_tmpl = config.get('costo_formula', '=+{col}{price_row}*{qty}')
    ct_row = rm['costo_total']
    for i, p in enumerate(provs):
        col = get_column_letter(3 + i)
        precio = p.get('precio')
        has_price = isinstance(precio, (int, float)) or (isinstance(precio, str) and precio.startswith('='))
        if has_price:
            f = ct_tmpl.format(col=col, price_row=rm['precio'], qty=qty, envio_row=rm['envio'])
            sc(ws[f'{col}{ct_row}'], f, FONT10B, NO_FILL, ACC, brd_ct, MONEY)
        else:
            sc(ws[f'{col}{ct_row}'], '-', FONT10, NO_FILL, ACC, brd_ct, MONEY)

    # Ahorro $$$
    ws.row_dimensions[rm['ahorro_d']].height = 20
    brd_ad = gbrd(rm['ahorro_d'])
    sc(ws[f'B{rm["ahorro_d"]}'], 'Ahorro en $$$ ', FONT10B, NO_FILL, ACC, brd_ad)
    for i, p in enumerate(provs):
        col = get_column_letter(3 + i)
        if i == 0:
            sc(ws[f'{col}{rm["ahorro_d"]}'], 'Referencia', FONT10, NO_FILL, ACC, brd_ad, MONEY)
        else:
            precio = p.get('precio')
            has_price = isinstance(precio, (int, float)) or (isinstance(precio, str) and str(precio).startswith('='))
            if has_price:
                sc(ws[f'{col}{rm["ahorro_d"]}'], f'=+$C${ct_row}-{col}{ct_row}', FONT10, NO_FILL, ACC, brd_ad, MONEY)
            else:
                sc(ws[f'{col}{rm["ahorro_d"]}'], '-', FONT10, NO_FILL, ACC, brd_ad, MONEY)

    # Ahorro %
    ws.row_dimensions[rm['ahorro_p']].height = 20
    brd_ap = gbrd(rm['ahorro_p'])
    sc(ws[f'B{rm["ahorro_p"]}'], 'Ahorro %', FONT10B, NO_FILL, ACC, brd_ap)
    for i, p in enumerate(provs):
        col = get_column_letter(3 + i)
        if i == 0:
            sc(ws[f'{col}{rm["ahorro_p"]}'], ' -', FONT10, NO_FILL, ACC, brd_ap, PCT)
        else:
            precio = p.get('precio')
            has_price = isinstance(precio, (int, float)) or (isinstance(precio, str) and str(precio).startswith('='))
            if has_price:
                sc(ws[f'{col}{rm["ahorro_p"]}'], f'=1-({col}{ct_row}/$C${ct_row})', FONT10, NO_FILL, ACC, brd_ap, PCT)
            else:
                sc(ws[f'{col}{rm["ahorro_p"]}'], '-', FONT10, NO_FILL, ACC, brd_ap, PCT)

    # CAPACIDAD Y LOGÍSTICA
    ws.row_dimensions[rm['cap_log']].height = 20
    sc(ws[f'B{rm["cap_log"]}'], 'CAPACIDAD Y LOGÍSTICA', FONT10B, NO_FILL, ACC, gbrd(rm['cap_log']))

    # Provincia
    ws.row_dimensions[rm['provincia']].height = 20
    sc(ws[f'B{rm["provincia"]}'], 'Provincia', FONT10B, NO_FILL, ACC, BN)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["provincia"]}'], p.get('provincia', '-'), FONT10, NO_FILL, ACC, BN)

    # Custom rows
    if custom:
        for ci, cr in enumerate(custom):
            cr_row = rm[f'custom_{ci}']
            ws.row_dimensions[cr_row].height = cr.get('height', 20)
            cr_font = FONT10B if cr.get('bold') else FONT10
            cr_align = ACCW if cr.get('wrap') else ACC
            sc(ws[f'B{cr_row}'], cr['label'], cr_font, NO_FILL, cr_align, BN)
            for i, p in enumerate(provs):
                vals = p.get('custom', [])
                val = vals[ci] if ci < len(vals) else '-'
                cr_nf = MONEY if cr.get('format') == 'money' else (PCT if cr.get('format') == 'pct' else 'General')
                sc(ws[f'{get_column_letter(3+i)}{cr_row}'], val, FONT10, NO_FILL, cr_align, BN, cr_nf)

    # Plazo
    ws.row_dimensions[rm['plazo']].height = 20
    brd_pl = gbrd(rm['plazo'])
    sc(ws[f'B{rm["plazo"]}'], 'Plazo de entrega (días)', FONT10, NO_FILL, ACC, brd_pl)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["plazo"]}'], p.get('plazo_entrega', '-'), FONT10, NO_FILL, ACC, brd_pl)

    # CONDICIONES COMERCIALES
    ws.row_dimensions[rm['cond_com']].height = 20
    sc(ws[f'B{rm["cond_com"]}'], 'CONDICIONES COMERCIALES', FONT10B, NO_FILL, ACC, gbrd(rm['cond_com']))

    # Financiación
    ws.row_dimensions[rm['financiacion']].height = config.get('financiacion_height', 44)
    brd_fin = gbrd(rm['financiacion'])
    sc(ws[f'B{rm["financiacion"]}'], 'Financiacion', FONT10, NO_FILL, ACC, brd_fin)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["financiacion"]}'], p.get('financiacion', '-'), FONT10, NO_FILL, ACCW, brd_fin)

    # SERVICIO AL CLIENTE
    ws.row_dimensions[rm['serv_cli']].height = 20
    sc(ws[f'B{rm["serv_cli"]}'], 'SERVICIO AL CLIENTE Y SOPORTE', FONT10B, NO_FILL, ACC, gbrd(rm['serv_cli']))

    # Garantía
    ws.row_dimensions[rm['garantia']].height = 20
    brd_g = gbrd(rm['garantia'])
    sc(ws[f'B{rm["garantia"]}'], 'Garantía del producto/servicio', FONT10, NO_FILL, ACC, brd_g)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["garantia"]}'], p.get('garantia', '-'), FONT10, NO_FILL, ACC, brd_g)

    # Proveedor Recomendado
    ws.row_dimensions[rm['recomendado']].height = 20
    brd_rec = gbrd(rm['recomendado'])
    sc(ws[f'B{rm["recomendado"]}'], 'Proveedor Recomendado BuyFlow', FONT10, PURPLE, ACC, brd_rec, PCT)
    for i, p in enumerate(provs):
        col = get_column_letter(3 + i)
        rv = p.get('recomendado')
        if rv:
            sc(ws[f'{col}{rm["recomendado"]}'], rv, FONT10, PURPLE, ACC, brd_rec, PCT)
        else:
            sc(ws[f'{col}{rm["recomendado"]}'], None, FONT10, NO_FILL, ACC, brd_rec)

    # Notas
    ws.row_dimensions[rm['notas']].height = config.get('notas_height', 80)
    brd_n = gbrd(rm['notas'])
    sc(ws[f'B{rm["notas"]}'], 'Notas y Observaciones Adicionales', FONT10, NO_FILL, ACCW, brd_n)
    for i, p in enumerate(provs):
        sc(ws[f'{get_column_letter(3+i)}{rm["notas"]}'], p.get('notas') or '-', FONT10, NO_FILL, ACCW, brd_n)

    # Logo
    if os.path.exists(LOGO_SRC):
        try:
            from PIL import Image as PILImg
            pil = PILImg.open(LOGO_SRC)
            new_h = 45; new_w = int(new_h * pil.width / pil.height)
            resized = '/home/claude/logo_resized.png'
            pil.resize((new_w, new_h), PILImg.LANCZOS).save(resized)
            img = Image(resized)
            img.anchor = ws.cell(row=2, column=min(last_ci, 5)).coordinate
            ws.add_image(img)
        except Exception:
            pass

    # Save → recalc → re-set widths
    wb.save(output_path)
    if os.path.exists(RECALC):
        subprocess.run(['python3', RECALC, output_path], capture_output=True)
        wb2 = load_workbook(output_path)
        ws2 = wb2.active
        ws2.column_dimensions['A'].width = 5.3
        ws2.column_dimensions['B'].width = 35
        for i in range(n):
            ws2.column_dimensions[get_column_letter(3 + i)].width = 28
        wb2.save(output_path)

    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════

def _is_purple(cell):
    fill = cell.fill
    if not fill or fill.fill_type not in ('solid',): return False
    try: return (fill.fgColor.rgb or '').upper() in ('FFB24EFF', 'B24EFF')
    except: return False

def _has_content(cell):
    return cell.value is not None and str(cell.value).strip() != ''

def _cref(row, col):
    return f"{get_column_letter(col)}{row}"

def _get_last_prov_col(ws):
    last = 3
    for col in range(3, ws.max_column + 1):
        if ws.cell(9, col).value is not None and str(ws.cell(9, col).value).strip():
            last = col
    return last

def _detect_flags(ws):
    has_desc = has_link = False
    for r in range(10, 15):
        val = str(ws.cell(r, 2).value or '').lower().strip()
        if val == 'descripcion': has_desc = True
        if val == 'link': has_link = True
    return has_desc, has_link

def _detect_custom_rows(ws, rm_base):
    provincia_row = rm_base['provincia']
    custom = []
    r = provincia_row + 1
    while r <= ws.max_row:
        val = str(ws.cell(r, 2).value or '').strip()
        if not val: r += 1; continue
        if 'plazo' in val.lower(): break
        is_bold = ws.cell(r, 2).font.bold if ws.cell(r, 2).font else False
        custom.append({"label": val, "bold": is_bold})
        r += 1
    return custom if custom else None


def validate_sheet(ws, tiene_desc=None, tiene_link=None):
    issues = []
    if tiene_desc is None or tiene_link is None:
        d, l = _detect_flags(ws)
        if tiene_desc is None: tiene_desc = d
        if tiene_link is None: tiene_link = l

    rm_base = row_map(tiene_desc, tiene_link, None)
    custom = _detect_custom_rows(ws, rm_base)
    rm = row_map(tiene_desc, tiene_link, custom)
    br = border_rules(rm)
    last_col = _get_last_prov_col(ws)
    prov_cols = list(range(4, last_col + 1))
    all_data_cols = list(range(3, last_col + 1))
    expected_last = rm['notas']
    mode_str = f"desc={'sí' if tiene_desc else 'no'} link={'sí' if tiene_link else 'no'}" + \
               (f" +{len(custom)} custom" if custom else "")

    # STRUCTURE
    for row in range(expected_last + 1, min(ws.max_row + 1, expected_last + 10)):
        for col in range(2, last_col + 1):
            if ws.cell(row, col).value is not None:
                issues.append(f"[ESTRUCTURA] Contenido en R{row} — layout debe terminar en R{expected_last}")
                break
    for row in range(1, expected_last + 1):
        if ws.cell(row, 1).value is not None:
            issues.append(f"[ESTRUCTURA] A{row}: columna A debe estar vacía")
    for row, col, substr, exp in [
        (rm['req'], 2, 'requerimiento', 'Requerimiento:'),
        (rm['rec'], 2, 'recomendaci', 'Recomendación BuyFlow'),
        (rm['prov_anal'], 2, 'proveedores', 'Proveedores analizados'),
    ]:
        val = ws.cell(row, col).value
        if not val or substr not in str(val).lower():
            issues.append(f"[ESTRUCTURA] {_cref(row,col)}: esperado '{exp}', encontrado: '{val}'")
    c9 = ws.cell(9, 3).value
    if not c9 or 'referencia' not in str(c9).lower():
        issues.append(f"[ESTRUCTURA] C9 debe ser 'Referencia', encontrado: '{c9}'")
    c7 = ws.cell(7, 3).value
    if c7 is None:
        issues.append("[ESTRUCTURA] C7 vacío")
    else:
        try: int(float(str(c7)))
        except: issues.append(f"[ESTRUCTURA] C7 debe ser número: '{c7}'")
    c_ahd = ws.cell(rm['ahorro_d'], 3).value
    if not c_ahd or 'referencia' not in str(c_ahd).lower():
        issues.append(f"[ESTRUCTURA] C{rm['ahorro_d']} debe ser 'Referencia': '{c_ahd}'")
    c_ahp = ws.cell(rm['ahorro_p'], 3).value
    if c_ahp is None or str(c_ahp).strip() not in ('-', ' -'):
        issues.append(f"[ESTRUCTURA] C{rm['ahorro_p']} debe ser '-': '{c_ahp}'")
    merged = [str(m) for m in ws.merged_cells.ranges]
    ll = get_column_letter(last_col)
    for r in [rm['req'], rm['rec']]:
        exp = f"C{r}:{ll}{r}"
        if exp not in merged:
            issues.append(f"[ESTRUCTURA] Merge esperado {exp} — encontrados: {merged}")

    # FILLS
    for row in [rm['req'], rm['rec'], rm['prov_anal']]:
        if not _is_purple(ws.cell(row, 2)):
            issues.append(f"[COLOR] B{row}: falta fill púrpura")
    for col in range(2, last_col + 1):
        if not _is_purple(ws.cell(rm['headers'], col)):
            issues.append(f"[COLOR] {_cref(rm['headers'], col)}: header sin fill púrpura")
    rr = rm['recomendado']
    if not _is_purple(ws.cell(rr, 2)):
        issues.append(f"[COLOR] B{rr}: falta fill púrpura")
    for col in prov_cols:
        cell = ws.cell(rr, col)
        val = str(cell.value or '').strip().upper()
        hp = _is_purple(cell)
        should = val in ('X', 'ALTERNATIVA')
        if should and not hp: issues.append(f"[COLOR] {_cref(rr, col)}: '{val}' sin fill púrpura")
        if not should and val == '' and hp: issues.append(f"[COLOR] {_cref(rr, col)}: vacía con fill (no debe)")

    # ALIGNMENT
    for row in range(1, expected_last + 1):
        for col in range(2, last_col + 1):
            cell = ws.cell(row, col)
            if not _has_content(cell): continue
            a = cell.alignment
            if (a.horizontal if a else None) != 'center' or (a.vertical if a else None) != 'center':
                issues.append(f"[ALINEACIÓN] {_cref(row, col)}: debe ser center/center")

    # GRIDLINES
    if ws.sheet_view.showGridLines:
        issues.append("[VISUAL] Gridlines visibles")

    # TYPOGRAPHY
    for row in range(1, expected_last + 1):
        esz = 14 if row in (rm['req'], rm['rec']) else 10
        for col in range(2, last_col + 1):
            cell = ws.cell(row, col)
            if not _has_content(cell): continue
            f = cell.font
            if f and f.name and f.name != 'Arial':
                issues.append(f"[TIPOGRAFÍA] {_cref(row, col)}: '{f.name}' → Arial")
            if f and f.size and f.size != esz:
                issues.append(f"[TIPOGRAFÍA] {_cref(row, col)}: {f.size} → {esz}")
    for row in bold_rows(rm):
        c = ws.cell(row, 2)
        if _has_content(c) and not (c.font and c.font.bold):
            issues.append(f"[TIPOGRAFÍA] B{row}: debe ser bold")
    for row in not_bold_rows(rm):
        c = ws.cell(row, 2)
        if _has_content(c) and c.font and c.font.bold:
            issues.append(f"[TIPOGRAFÍA] B{row}: NO debe ser bold")
    for row in [rm['costo_total'], rm['ahorro_d'], rm['ahorro_p']]:
        for col in prov_cols:
            c = ws.cell(row, col)
            if _has_content(c) and not (c.font and c.font.bold):
                issues.append(f"[TIPOGRAFÍA] {_cref(row, col)}: valores de ahorro/costo deben ser bold")

    # FORMULAS
    ct = rm['costo_total']
    for col in prov_cols:
        v = str(ws.cell(ct, col).value or '').strip()
        if v not in ('', '-', 'A definir') and not v.startswith('='):
            issues.append(f"[FÓRMULAS] {_cref(ct, col)}: Costo Total hardcodeado")
    for col in prov_cols:
        v = str(ws.cell(rm['ahorro_d'], col).value or '').strip()
        if v in ('', '-'): continue
        if not v.startswith('='): issues.append(f"[FÓRMULAS] {_cref(rm['ahorro_d'], col)}: Ahorro$ hardcodeado")
        elif f'$C${ct}' not in v.upper(): issues.append(f"[FÓRMULAS] {_cref(rm['ahorro_d'], col)}: no ref $C${ct}")
    for col in prov_cols:
        v = str(ws.cell(rm['ahorro_p'], col).value or '').strip()
        if v in ('', '-'): continue
        if not v.startswith('='): issues.append(f"[FÓRMULAS] {_cref(rm['ahorro_p'], col)}: Ahorro% hardcodeado")
        elif f'$C${ct}' not in v.upper(): issues.append(f"[FÓRMULAS] {_cref(rm['ahorro_p'], col)}: no ref $C${ct}")

    # NUMBER FORMATS
    for row in [rm['precio'], rm['costo_total'], rm['ahorro_d']]:
        for col in all_data_cols:
            cell = ws.cell(row, col)
            if not _has_content(cell): continue
            v = str(cell.value).strip()
            if v in ('-', 'Referencia', 'A definir', ' -'): continue
            nf = cell.number_format or ''
            if '$' not in nf and (v.startswith('=') or isinstance(cell.value, (int, float))):
                issues.append(f"[FORMATO] {_cref(row, col)}: falta formato moneda $")
    for col in prov_cols:
        cell = ws.cell(rm['ahorro_p'], col)
        if not _has_content(cell): continue
        v = str(cell.value).strip()
        if v in ('-', ' -', ''): continue
        if '%' not in (cell.number_format or ''):
            issues.append(f"[FORMATO] {_cref(rm['ahorro_p'], col)}: falta formato %")

    # BORDERS
    for row, sides in br.items():
        for col in range(2, last_col + 1):
            cell = ws.cell(row, col)
            if not _has_content(cell) and col != 2: continue
            b = cell.border
            for side in sides:
                s = getattr(b, side, None) if b else None
                if (s.style if s else None) != 'medium':
                    issues.append(f"[BORDES] {_cref(row, col)}: {side} debe ser 'medium'")

    # COLUMN WIDTHS
    bw = ws.column_dimensions['B'].width
    if bw and not (30 <= bw <= 40):
        issues.append(f"[LAYOUT] Col B ancho {bw:.1f} → ~35")
    for col in range(3, last_col + 1):
        w = ws.column_dimensions[get_column_letter(col)].width
        if w and not (24 <= w <= 34):
            issues.append(f"[LAYOUT] Col {get_column_letter(col)} ancho {w:.1f} → 28-30")

    # RECOMMENDATION
    c5 = ws.cell(rm['rec'], 3).value
    if not c5:
        issues.append("[CONTENIDO] Recomendación vacía")
    else:
        if len(str(c5).split()) > 100:
            issues.append("[CONTENIDO] Recomendación muy larga")
        if 'recom' not in str(c5).lower():
            issues.append("[CONTENIDO] Recomendación no menciona 'recomienda'")

    # WRAP TEXT
    wr = [rm['req'], rm['rec'], rm['notas']]
    if 'descripcion' in rm: wr.append(rm['descripcion'])
    for row in wr:
        for col in range(3, last_col + 1):
            cell = ws.cell(row, col)
            if _has_content(cell) and cell.alignment and not cell.alignment.wrap_text:
                issues.append(f"[FORMATO] {_cref(row, col)}: falta wrap_text")

    return issues, mode_str


def validate_file(filepath, tiene_desc=None, tiene_link=None):
    print(f"\n{'═'*62}")
    print(f"  BuyFlow Validator v3.0")
    print(f"  Archivo: {filepath}")
    print(f"{'═'*62}\n")
    try: wb = load_workbook(filepath, data_only=False)
    except Exception as e: print(f"  ERROR: {e}\n"); return False

    total = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  Hoja: '{sn}'")
        print(f"  {'─'*56}")
        issues, mode_str = validate_sheet(ws, tiene_desc, tiene_link)
        print(f"  Modo: {mode_str}")
        total += len(issues)
        if not issues:
            print("  PASS\n")
        else:
            print(f"  FAIL — {len(issues)} error(es):\n")
            by_cat = {}
            for iss in issues:
                cat = iss.split(']')[0].replace('[', '') if ']' in iss else 'OTRO'
                by_cat.setdefault(cat, []).append(iss)
            for cat, items in by_cat.items():
                print(f"     [{cat}]")
                for it in items:
                    print(f"       - {it.split(']', 1)[-1].strip()}")
            print()
    print('═' * 62)
    print(f"  {'VÁLIDO' if total == 0 else f'INVÁLIDO — {total} error(es)'}")
    print('═' * 62 + '\n')
    return total == 0


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("python3 buyflow.py validate <reporte.xlsx>")
        print("python3 buyflow.py generate <config.json> [output.xlsx]")
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == 'validate':
        sys.exit(0 if validate_file(sys.argv[2]) else 1)
    elif cmd == 'generate':
        with open(sys.argv[2]) as f: cfg = json.load(f)
        out = sys.argv[3] if len(sys.argv) > 3 else '/home/claude/report.xlsx'
        generate_report(cfg, out)
        sys.exit(0 if validate_file(out) else 1)
